import openpyxl as xl
import os
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
import time

# Update the current working directory to the main project folder
os.chdir(os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))

# Read the trivia list data
trivia_wb = xl.load_workbook('data/cbus_trivia_list.xlsx')
sheet = trivia_wb.active

# Set up Selenium
options = Options()
options.add_argument('--headless')
driver = webdriver.Chrome(options=options)

# Create an additional column to extract the street address
link_cols = [1, 6, 11, 16, 21]

for col in reversed(link_cols):
    street_idx = col + 4
    sheet.insert_cols(street_idx)
    sheet.cell(row=1, column=street_idx).value = 'Address'

    num_empty_rows = 0
    custom_max_row = sheet.max_row
    for row in range(2, sheet.max_row + 1):
        source_cell = sheet.cell(row=row, column=col)

        if not source_cell.value:
           num_empty_rows += 1
        else:
           num_empty_rows = 0

        if num_empty_rows >=4:
            custom_max_row = row - 4
            break

    for row in range(2, custom_max_row+1):
        source_cell = sheet.cell(row=row, column=col)

        if source_cell.hyperlink:
            driver.get(source_cell.hyperlink.target)
            time.sleep(2)
            try:
                address = driver.find_element(By.XPATH,
                               "//button[(@data-item-id='address')]")
                raw_address = address.get_attribute("aria-label")
                sheet.cell(row=row, column=street_idx).value = raw_address[8:]
            except Exception:
                sheet.cell(row=row, column=street_idx).value = None


# Save the new columns in a new workbook
trivia_wb.save('data/cbus_trivia_list_updated.xlsx')